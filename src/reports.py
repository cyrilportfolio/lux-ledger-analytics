"""Statements produced from a ledger once it has passed the checks.

Trial balance, customer and supplier balances, balance sheet and profit and
loss account mapped onto the PCN 2020 "tableau de passage", VAT return
worked out rate by rate, and a formatted Excel workbook holding all of it.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd

from src import config


# --------------------------------------------------------------------------
# Reference
# --------------------------------------------------------------------------
def load_postes(path: Path | None = None) -> pd.DataFrame:
    """Ordered list of balance sheet and profit and loss captions."""
    path = Path(path or config.DATA_DIR / "postes.csv")
    frame = pd.read_csv(path, sep=config.CSV_SEPARATOR, dtype=str,
                        keep_default_na=False, encoding="utf-8")
    frame.columns = [c.strip().lower() for c in frame.columns]
    frame["ordre"] = pd.to_numeric(frame["ordre"], errors="coerce").fillna(999)
    for column in ("rubrique", "poste", "sens"):
        frame[column] = frame[column].str.strip()
    return frame


# --------------------------------------------------------------------------
# Balances
# --------------------------------------------------------------------------
def trial_balance(journal: pd.DataFrame, pcn: pd.DataFrame) -> pd.DataFrame:
    """General trial balance: totals and closing balance per account."""
    grouped = (journal.groupby("compte", as_index=False)
               .agg(debit=("debit", "sum"), credit=("credit", "sum"),
                    mouvements=("compte", "size")))
    grouped[["debit", "credit"]] = grouped[["debit", "credit"]].round(2)
    grouped["solde"] = (grouped["debit"] - grouped["credit"]).round(2)
    grouped["solde_debiteur"] = grouped["solde"].clip(lower=0)
    grouped["solde_crediteur"] = (-grouped["solde"]).clip(lower=0)

    reference = pcn[["compte", "libelle", "classe", "rubrique", "poste"]]
    merged = grouped.merge(reference, on="compte", how="left")
    merged["libelle"] = merged["libelle"].fillna("COMPTE HORS PCN")
    merged["classe"] = merged["classe"].fillna(merged["compte"].str[0])
    columns = ["compte", "libelle", "classe", "mouvements", "debit", "credit",
               "solde_debiteur", "solde_crediteur"]
    return merged.sort_values("compte")[columns].reset_index(drop=True)


def auxiliary_balance(journal: pd.DataFrame, prefix: str, label: str,
                      third_parties: pd.DataFrame | None = None) -> pd.DataFrame:
    """Open items per customer or per supplier."""
    subset = journal.loc[journal["compte"].str.startswith(prefix)].copy()
    subset = subset.loc[subset["tiers"].fillna("").ne("")]
    if subset.empty:
        return pd.DataFrame(columns=["tiers", "nom", "debit", "credit", "solde",
                                     "derniere_ecriture"])
    grouped = (subset.groupby("tiers", as_index=False)
               .agg(debit=("debit", "sum"), credit=("credit", "sum"),
                    derniere_ecriture=("date", "max")))
    grouped[["debit", "credit"]] = grouped[["debit", "credit"]].round(2)
    grouped["solde"] = (grouped["debit"] - grouped["credit"]).round(2)
    grouped["type"] = label
    if third_parties is not None and not third_parties.empty:
        grouped = grouped.merge(third_parties[["code", "nom", "pays", "numero_tva"]],
                                left_on="tiers", right_on="code", how="left")
        grouped = grouped.drop(columns=["code"])
    else:
        grouped["nom"] = ""
    grouped["derniere_ecriture"] = pd.to_datetime(grouped["derniere_ecriture"])
    columns = [c for c in ["tiers", "nom", "pays", "numero_tva", "type", "debit",
                           "credit", "solde", "derniere_ecriture"]
               if c in grouped.columns]
    return grouped.sort_values("solde", ascending=False)[columns].reset_index(drop=True)


def _aggregate_by_poste(journal: pd.DataFrame, postes: pd.DataFrame,
                        rubrique: str, signed: pd.Series) -> pd.DataFrame:
    subset = journal.loc[journal["rubrique"].fillna("").eq(rubrique)].copy()
    subset["valeur"] = signed.loc[subset.index]
    grouped = subset.groupby("poste", as_index=False)["valeur"].sum()
    reference = postes.loc[postes["rubrique"].eq(rubrique)]
    merged = reference.merge(grouped, on="poste", how="left")
    merged["valeur"] = merged["valeur"].fillna(0.0).round(2)
    return merged.sort_values("ordre")[["poste", "valeur"]].reset_index(drop=True)


def income_statement(journal: pd.DataFrame, postes: pd.DataFrame) -> pd.DataFrame:
    """Profit and loss account in list form, PCN captions, in order."""
    subset = journal.loc[journal["rubrique"].fillna("").eq("CPP")].copy()
    if subset.empty:
        return pd.DataFrame(columns=["poste", "montant"])
    subset["mouvement"] = (subset["credit"] - subset["debit"]).round(2)
    grouped = subset.groupby("poste", as_index=False)["mouvement"].sum()

    reference = postes.loc[postes["rubrique"].eq("CPP")]
    merged = reference.merge(grouped, on="poste", how="left")
    merged["mouvement"] = merged["mouvement"].fillna(0.0)
    # Expense captions are shown positive, the way the eCDF layout reads them.
    merged["montant"] = np.where(merged["sens"].eq("-"),
                                 -merged["mouvement"], merged["mouvement"]).round(2)
    result = merged.sort_values("ordre")[["poste", "sens", "montant"]]
    total = round(float(merged["mouvement"].sum()), 2)
    result = pd.concat([
        result,
        pd.DataFrame([{"poste": "17. Resultat de l'exercice", "sens": "=",
                       "montant": total}]),
    ], ignore_index=True)
    return result


def net_result(journal: pd.DataFrame) -> float:
    """Profit or loss of the period, from classes 6 and 7."""
    subset = journal.loc[journal["classe"].isin(["6", "7"])]
    return round(float((subset["credit"] - subset["debit"]).sum()), 2)


def balance_sheet(journal: pd.DataFrame, postes: pd.DataFrame) -> pd.DataFrame:
    """Balance sheet by PCN caption, with the result of the period included."""
    assets_signed = (journal["debit"] - journal["credit"]).round(2)
    liabilities_signed = (journal["credit"] - journal["debit"]).round(2)

    assets = _aggregate_by_poste(journal, postes, "BILAN_ACTIF", assets_signed)
    assets["cote"] = "ACTIF"
    liabilities = _aggregate_by_poste(journal, postes, "BILAN_PASSIF",
                                      liabilities_signed)
    liabilities["cote"] = "PASSIF"

    result = net_result(journal)
    mask = liabilities["poste"].eq("A.VI. Resultat de l'exercice")
    liabilities.loc[mask, "valeur"] = (liabilities.loc[mask, "valeur"] + result).round(2)

    frame = pd.concat([assets, liabilities], ignore_index=True)
    frame = frame.rename(columns={"valeur": "montant"})
    totals = frame.groupby("cote")["montant"].sum().round(2)
    frame = pd.concat([
        frame,
        pd.DataFrame([
            {"cote": "ACTIF", "poste": "TOTAL DE L'ACTIF",
             "montant": float(totals.get("ACTIF", 0.0))},
            {"cote": "PASSIF", "poste": "TOTAL DU PASSIF",
             "montant": float(totals.get("PASSIF", 0.0))},
        ]),
    ], ignore_index=True)
    return frame[["cote", "poste", "montant"]]


def balance_sheet_control(bilan: pd.DataFrame) -> dict:
    """Assets minus liabilities: must be nil once the result is posted."""
    total_actif = float(bilan.loc[bilan["poste"].eq("TOTAL DE L'ACTIF"),
                                  "montant"].sum())
    total_passif = float(bilan.loc[bilan["poste"].eq("TOTAL DU PASSIF"),
                                   "montant"].sum())
    return {
        "total_actif": round(total_actif, 2),
        "total_passif": round(total_passif, 2),
        "ecart": round(total_actif - total_passif, 2),
        "equilibre": abs(total_actif - total_passif) <= config.BALANCE_TOLERANCE,
    }


# --------------------------------------------------------------------------
# VAT
# --------------------------------------------------------------------------
def vat_detail(journal: pd.DataFrame, vat_codes: pd.DataFrame) -> pd.DataFrame:
    """Taxable bases and VAT, per period and per VAT code."""
    rates = vat_codes.set_index("code")
    taxed = journal.loc[journal["code_tva"].isin(rates.index)].copy()
    taxed = taxed.loc[taxed["code_tva"].ne("NA")]
    if taxed.empty:
        return pd.DataFrame(columns=["periode", "code_tva", "libelle", "sens",
                                     "taux", "base", "taxe_theorique"])
    taxed["sens"] = taxed["code_tva"].map(rates["sens"])
    taxed["taux"] = taxed["code_tva"].map(rates["taux"])
    taxed["libelle"] = taxed["code_tva"].map(rates["libelle"])
    taxed["base"] = np.where(taxed["sens"].eq("VENTE"),
                             taxed["credit"] - taxed["debit"],
                             taxed["debit"] - taxed["credit"])
    grouped = (taxed.groupby(["periode", "code_tva", "libelle", "sens", "taux"],
                             as_index=False)["base"].sum())
    grouped["base"] = grouped["base"].round(2)
    grouped["taxe_theorique"] = (grouped["base"] * grouped["taux"]).round(2)
    return grouped.sort_values(["periode", "sens", "code_tva"]).reset_index(drop=True)


def vat_return(detail: pd.DataFrame) -> pd.DataFrame:
    """Draft VAT return for the year: bases by rate, output and input tax."""
    if detail.empty:
        return pd.DataFrame(columns=["rubrique", "libelle", "base", "taxe"])
    annual = (detail.groupby(["sens", "code_tva", "libelle", "taux"],
                             as_index=False)
              .agg(base=("base", "sum"), taxe=("taxe_theorique", "sum")))
    annual[["base", "taxe"]] = annual[["base", "taxe"]].round(2)

    rows = []
    sales = annual.loc[annual["sens"].eq("VENTE")].sort_values("taux",
                                                               ascending=False)
    for _, row in sales.iterrows():
        caption = (f"Operations imposables au taux de {row['taux'] * 100:.0f} %"
                   if row["taux"] > 0 else f"Operations exonerees - {row['libelle']}")
        rows.append({"rubrique": "TVA COLLECTEE", "libelle": caption,
                     "code": row["code_tva"], "base": row["base"],
                     "taxe": row["taxe"]})
    collected = round(float(sales["taxe"].sum()), 2)
    rows.append({"rubrique": "TVA COLLECTEE", "libelle": "Total de la taxe en aval",
                 "code": "", "base": round(float(sales["base"].sum()), 2),
                 "taxe": collected})

    purchases = annual.loc[annual["sens"].eq("ACHAT")].sort_values("taux",
                                                                   ascending=False)
    for _, row in purchases.iterrows():
        caption = (f"Achats au taux de {row['taux'] * 100:.0f} %"
                   if row["taux"] > 0 else f"Achats - {row['libelle']}")
        rows.append({"rubrique": "TVA DEDUCTIBLE", "libelle": caption,
                     "code": row["code_tva"], "base": row["base"],
                     "taxe": row["taxe"]})
    deductible = round(float(purchases["taxe"].sum()), 2)
    rows.append({"rubrique": "TVA DEDUCTIBLE",
                 "libelle": "Total de la taxe en amont", "code": "",
                 "base": round(float(purchases["base"].sum()), 2),
                 "taxe": deductible})

    net = round(collected - deductible, 2)
    rows.append({"rubrique": "SOLDE",
                 "libelle": "TVA due" if net >= 0 else "Credit de TVA",
                 "code": "", "base": np.nan, "taxe": abs(net)})
    return pd.DataFrame(rows)[["rubrique", "libelle", "code", "base", "taxe"]]


def vat_monthly(detail: pd.DataFrame) -> pd.DataFrame:
    """Month by month view: output tax, input tax, balance."""
    if detail.empty:
        return pd.DataFrame(columns=["periode", "taxe_collectee",
                                     "taxe_deductible", "solde"])
    pivot = (detail.pivot_table(index="periode", columns="sens",
                                values="taxe_theorique", aggfunc="sum")
             .fillna(0.0).round(2).reset_index())
    for column in ("VENTE", "ACHAT"):
        if column not in pivot.columns:
            pivot[column] = 0.0
    pivot = pivot.rename(columns={"VENTE": "taxe_collectee",
                                  "ACHAT": "taxe_deductible"})
    pivot["solde"] = (pivot["taxe_collectee"] - pivot["taxe_deductible"]).round(2)
    return pivot[["periode", "taxe_collectee", "taxe_deductible", "solde"]]


# --------------------------------------------------------------------------
# Journal-level views
# --------------------------------------------------------------------------
def journal_summary(journal: pd.DataFrame) -> pd.DataFrame:
    """Totals per journal and per period, the first thing a reviewer opens."""
    grouped = (journal.groupby(["journal", "periode"], as_index=False)
               .agg(pieces=("piece", "nunique"), lignes=("piece", "size"),
                    debit=("debit", "sum"), credit=("credit", "sum")))
    grouped[["debit", "credit"]] = grouped[["debit", "credit"]].round(2)
    grouped["ecart"] = (grouped["debit"] - grouped["credit"]).round(2)
    return grouped.sort_values(["journal", "periode"]).reset_index(drop=True)


def build_all(journal: pd.DataFrame, pcn: pd.DataFrame, vat_codes: pd.DataFrame,
              postes: pd.DataFrame,
              third_parties: pd.DataFrame | None = None) -> dict:
    """Every statement, in one dictionary keyed by sheet name."""
    detail = vat_detail(journal, vat_codes)
    bilan = balance_sheet(journal, postes)
    return {
        "journaux": journal_summary(journal),
        "balance_generale": trial_balance(journal, pcn),
        "balance_clients": auxiliary_balance(
            journal, config.CUSTOMER_ACCOUNT_PREFIX, "CLIENT", third_parties),
        "balance_fournisseurs": auxiliary_balance(
            journal, config.SUPPLIER_ACCOUNT_PREFIX, "FOURNISSEUR", third_parties),
        "bilan": bilan,
        "compte_de_resultat": income_statement(journal, postes),
        "tva_declaration": vat_return(detail),
        "tva_par_mois": vat_monthly(detail),
        "tva_detail": detail,
        "_controle_bilan": balance_sheet_control(bilan),
    }


# --------------------------------------------------------------------------
# Excel export
# --------------------------------------------------------------------------
SHEET_TITLES = {
    "synthese": "Synthese",
    "anomalies": "Anomalies",
    "controles": "Controles",
    "journaux": "Journaux",
    "balance_generale": "Balance generale",
    "balance_clients": "Balance clients",
    "balance_fournisseurs": "Balance fourn.",
    "bilan": "Bilan",
    "compte_de_resultat": "Compte de resultat",
    "tva_declaration": "TVA declaration",
    "tva_par_mois": "TVA par mois",
    "tva_detail": "TVA detail",
}

MONEY_COLUMNS = {"debit", "credit", "solde", "solde_debiteur", "solde_crediteur",
                 "montant", "base", "taxe", "taxe_theorique", "taxe_collectee",
                 "taxe_deductible", "ecart", "valeur"}


def _autosize(worksheet, frame: pd.DataFrame) -> None:
    from openpyxl.utils import get_column_letter

    for index, column in enumerate(frame.columns, start=1):
        longest = frame[column].astype(str).str.len().max()
        longest = 0 if pd.isna(longest) else int(longest)
        width = max(len(str(column)), longest) + 3
        worksheet.column_dimensions[get_column_letter(index)].width = min(width, 58)


def _style_sheet(worksheet, frame: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    worksheet.freeze_panes = "A2"
    if len(frame) and len(frame.columns):
        worksheet.auto_filter.ref = worksheet.dimensions

    # Printing: landscape, one page wide, header repeated on every page.
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = "1:1"

    for index, column in enumerate(frame.columns, start=1):
        if column in MONEY_COLUMNS:
            for row in range(2, len(frame) + 2):
                cell = worksheet.cell(row=row, column=index)
                cell.number_format = "#,##0.00"
        for row in range(2, len(frame) + 2):
            worksheet.cell(row=row, column=index).border = Border(bottom=thin)
    _autosize(worksheet, frame)


def _synthesis_frame(context: dict) -> pd.DataFrame:
    rows = [
        ("Entite", context["company"]),
        ("Numero de TVA", context["vat_number"]),
        ("Exercice", f"{context['start']:%d/%m/%Y} - {context['end']:%d/%m/%Y}"),
        ("Fichier source", context["source"]),
        ("Date d'execution", context["run_date"]),
        ("Lignes lues", context["rows_read"]),
        ("Lignes retenues", context["rows_kept"]),
        ("Pieces", context["pieces"]),
        ("Total debit", context["total_debit"]),
        ("Total credit", context["total_credit"]),
        ("Ecart debit/credit", context["ecart"]),
        ("Total de l'actif", context["total_actif"]),
        ("Total du passif", context["total_passif"]),
        ("Resultat de l'exercice", context["resultat"]),
        ("TVA nette de l'exercice", context["tva_nette"]),
        ("Anomalies detectees", context["anomalies"]),
        ("dont bloquantes", context["anomalies_bloquantes"]),
    ]
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def write_workbook(path: Path, reports: dict, anomalies: pd.DataFrame,
                   summary: pd.DataFrame, context: dict) -> Path:
    """Write every statement into one formatted workbook."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("synthese", _synthesis_frame(context)),
        ("controles", summary),
        ("anomalies", anomalies),
    ]
    for key in ("journaux", "balance_generale", "balance_clients",
                "balance_fournisseurs", "bilan", "compte_de_resultat",
                "tva_declaration", "tva_par_mois", "tva_detail"):
        sheets.append((key, reports[key]))

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for key, frame in sheets:
            title = SHEET_TITLES.get(key, key)[:31]
            export = frame.copy()
            for column in export.columns:
                if pd.api.types.is_datetime64_any_dtype(export[column]):
                    export[column] = export[column].dt.strftime("%d/%m/%Y")
            export.to_excel(writer, sheet_name=title, index=False)
            _style_sheet(writer.sheets[title], export)
    return path


def write_anomaly_report(path: Path, anomalies: pd.DataFrame,
                         summary: pd.DataFrame, context: dict) -> Path:
    """Plain text working list, the version that gets printed and ticked off."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "RAPPORT D'ANOMALIES",
        "=" * 72,
        f"Entite            : {context['company']}",
        f"Exercice          : {context['start']:%d/%m/%Y} - {context['end']:%d/%m/%Y}",
        f"Fichier analyse   : {context['source']}",
        f"Execution         : {context['run_date']}",
        f"Lignes analysees  : {context['rows_kept']}",
        f"Anomalies         : {context['anomalies']} "
        f"(dont {context['anomalies_bloquantes']} bloquantes)",
        "",
        "SYNTHESE DES CONTROLES",
        "-" * 72,
    ]
    for _, row in summary.iterrows():
        lines.append(f"{row['statut']:<11} {row['severite']:<9} "
                     f"{row['anomalies']:>5}  {row['libelle_controle']}")

    lines += ["", "DETAIL", "-" * 72]
    if anomalies.empty:
        lines.append("Aucune anomalie detectee.")
    else:
        for code, group in anomalies.groupby("code_controle", sort=False):
            lines.append("")
            lines.append(f"[{code}] {group.iloc[0]['libelle_controle']} "
                         f"- {len(group)} occurrence(s), "
                         f"severite {group.iloc[0]['severite']}")
            for _, row in group.head(50).iterrows():
                when = ("" if pd.isna(row["date"])
                        else pd.Timestamp(row["date"]).strftime("%d/%m/%Y"))
                lines.append(f"  {row['piece']:<14} {when:<11} "
                             f"{str(row['compte']):<8} {row['message']}")
            if len(group) > 50:
                lines.append(f"  ... {len(group) - 50} autre(s) occurrence(s)")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path
